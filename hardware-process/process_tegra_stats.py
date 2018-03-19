#!/usr/bin/env python3
import argparse
from os import listdir
from sys import maxsize
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter


parser = argparse.ArgumentParser(description="will output all the data for a given group test into an excel sheet")
parser.add_argument('--dir', metavar='store', type=str, help='the directory with the datawithin')
parser.add_argument('--output', metavar='store', type=str, help='where to save the value to (without extension)')

args = parser.parse_args()

def getDirectories(directory):
    return listdir(directory)

def processCPU(line):
    line = line.replace("[","")
    line = line.replace("]","")
    count = 0
    for i in line.split(","):
        i = i.split("@")
        count = count + int(i[0].replace("%", ""))
    return count


def processLine(line):
    l = line.rstrip()
    l = l.split(" ")
    if(len(l) > 1):
        return [l[1] , processCPU(l[9])]
    return ["0/0", 0]

def processFile(f):
    maxCPU = -1
    minCPU = 900000
    countCpu = 0
    count = 0
    maxRam = -1
    minRam = 9000000
    countRam = 0

    for line in f.readlines():
        processed = processLine(line)
        line = f.readline()
        count = count+1
        if maxCPU < processed[1]:
            maxCPU = processed[1]
        if minCPU > processed[1]:
            minCPU = processed[1]
        if minRam > int(processed[0].split('/')[0]):
            minRam = int(processed[0].split('/')[0])
        if maxRam < int(processed[0].split('/')[0]):
            maxRam = int(processed[0].split('/')[0])
        countRam = countRam + int(processed[0].split('/')[0])
        countCpu = countCpu + processed[1]
    values = [countCpu/count, maxCPU, minCPU, countRam/count, maxRam, minRam]
    return values
        

dirs = getDirectories(args.dir)

count = 0
wb = Workbook()

dest_filename = args.output + ".xlsx"

ws1 = wb.active
ws1.title = "Data"

rowCount = 7

places = dict()
counts = dict()
for i in dirs:
    f = args.dir + "/" + str(i)
    record = open(f, "r")
    if(len(record.readlines()) > 0):
        if str(i[0:-2]) not in places:
            places[str(i[0:-2])] = count
            counts[count] = 0
            ws1.cell(column=1, row = (count*rowCount)+1, value="Name")
            ws1.cell(column=2, row = (count*rowCount)+1, value=str(i[0:-2]))
            ws1.cell(column=1, row = (count*rowCount)+2, value="CPU Avg")
            ws1.cell(column=1, row = (count*rowCount)+3, value="Max CPU")
            ws1.cell(column=1, row = (count*rowCount)+4, value="Min CPU")
            ws1.cell(column=1, row = (count*rowCount)+5, value="Ram Avg")
            ws1.cell(column=1, row = (count*rowCount)+6, value="Max RAM")
            ws1.cell(column=1, row = (count*rowCount)+7, value="Min RAM")
            count = count + 1
        else:
            counts[places[str(i[0:-2])]] = counts[places[str(i[0:-2])]] + 1;
        record.seek(0)
        vals = processFile(record)
        for j in range(1,rowCount):
            ws1.cell(column=2 + counts[places[str(i[0:-2])]], row = (places[str(i[0:-2])]*rowCount) + j + 1, value=vals[j-1])

wb.save(filename = dest_filename)
